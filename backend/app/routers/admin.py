from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from ..database import get_db, row, rows
from ..deps import admin_user, current_user
from ..security import hash_password

router = APIRouter(prefix="/admin", tags=["admin"])


# ── GET all users ──────────────────────────────────────────────────────────
@router.get("/users")
def list_users(admin: dict = Depends(admin_user)):
    with get_db() as db:
        return rows(db.execute(
            """
            SELECT u.*,
                   (SELECT COUNT(*) FROM predictions p WHERE p.user_id = u.id) AS predictions_count
            FROM users u
            ORDER BY u.created_at DESC
            """
        ))


# ── Toggle user active/inactive ───────────────────────────────────────────
@router.put("/users/{user_id}/toggle-active")
def toggle_active(user_id: int, admin: dict = Depends(admin_user)):
    with get_db() as db:
        user = row(db.execute("SELECT id, is_active, name FROM users WHERE id = ?", (user_id,)))
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        new_status = 0 if user["is_active"] else 1
        db.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
        return {"id": user_id, "is_active": bool(new_status), "name": user["name"]}


# ── Delete user + all their predictions ───────────────────────────────────
@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin: dict = Depends(admin_user)):
    with get_db() as db:
        user = row(db.execute("SELECT id, name, role FROM users WHERE id = ?", (user_id,)))
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        if user["role"] == "admin":
            raise HTTPException(status_code=403, detail="Cannot delete an admin user")
        if user_id == admin["id"]:
            raise HTTPException(status_code=403, detail="Cannot delete yourself")
        db.execute("DELETE FROM predictions WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM leaderboards WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        return {"status": "deleted", "user_id": user_id, "name": user["name"]}


# ── Reset user password ───────────────────────────────────────────────────
class PasswordReset(BaseModel):
    password: str

@router.put("/users/{user_id}/reset-password")
def reset_password(user_id: int, payload: PasswordReset, admin: dict = Depends(admin_user)):
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    with get_db() as db:
        user = row(db.execute("SELECT id, name FROM users WHERE id = ?", (user_id,)))
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        db.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (hash_password(payload.password), user_id)
        )
        return {"status": "password_reset", "user_id": user_id, "name": user["name"]}


# ── DATA RESET ENDPOINTS ──────────────────────────────────────────────────
@router.delete("/reset/predictions")
def reset_predictions(admin: dict = Depends(admin_user)):
    with get_db() as db:
        res = db.execute("SELECT COUNT(*) AS cnt FROM predictions").fetchone()
        count = list(res.values())[0] if res else 0
        db.execute("DELETE FROM predictions")
        db.execute("UPDATE leaderboards SET total_points=0, exact_matches=0, winner_count=0, predictions_count=0, accuracy=0")
        return {"message": f"Cleared {count} predictions and reset leaderboard.", "deleted": count}


@router.delete("/reset/matches")
def reset_matches(admin: dict = Depends(admin_user)):
    with get_db() as db:
        res_pred = db.execute("SELECT COUNT(*) FROM predictions").fetchone()
        res_match = db.execute("SELECT COUNT(*) FROM matches").fetchone()
        res_team = db.execute("SELECT COUNT(*) FROM teams").fetchone()
        pred_count  = list(res_pred.values())[0] if res_pred else 0
        match_count = list(res_match.values())[0] if res_match else 0
        team_count  = list(res_team.values())[0] if res_team else 0
        db.execute("DELETE FROM predictions")
        db.execute("DELETE FROM leaderboards")
        db.execute("DELETE FROM match_history")
        db.execute("DELETE FROM matches")
        db.execute("DELETE FROM teams")
        return {
            "message": f"Cleared {match_count} matches, {team_count} teams, {pred_count} predictions.",
            "matches": match_count, "teams": team_count, "predictions": pred_count,
        }


@router.delete("/reset/tournaments")
def reset_tournaments(admin: dict = Depends(admin_user)):
    with get_db() as db:
        res_tourney = db.execute("SELECT COUNT(*) FROM tournaments").fetchone()
        res_match = db.execute("SELECT COUNT(*) FROM matches").fetchone()
        res_pred = db.execute("SELECT COUNT(*) FROM predictions").fetchone()
        counts = {
            "tournaments": list(res_tourney.values())[0] if res_tourney else 0,
            "matches":     list(res_match.values())[0] if res_match else 0,
            "predictions": list(res_pred.values())[0] if res_pred else 0,
        }
        db.execute("DELETE FROM predictions")
        db.execute("DELETE FROM leaderboards")
        db.execute("DELETE FROM match_history")
        db.execute("DELETE FROM matches")
        db.execute("DELETE FROM teams")
        db.execute("DELETE FROM tournaments")
        return {"message": "All tournament data cleared. Users kept.", **counts}


@router.delete("/reset/leaderboard")
def reset_leaderboard(admin: dict = Depends(admin_user)):
    with get_db() as db:
        res = db.execute("SELECT COUNT(*) FROM leaderboards").fetchone()
        count = list(res.values())[0] if res else 0
        db.execute("UPDATE leaderboards SET total_points=0, exact_matches=0, winner_count=0, accuracy=0, rank=NULL, badges=''")
        return {"message": f"Reset {count} leaderboard entries.", "reset": count}


# ── Analytics ──────────────────────────────────────────────────────────────
@router.get("/analytics")
def analytics(admin: dict = Depends(admin_user)):
    with get_db() as db:
        res_users     = db.execute("SELECT COUNT(*) FROM users WHERE role='user'").fetchone()
        res_matches   = db.execute("SELECT COUNT(*) FROM matches").fetchone()
        res_preds     = db.execute("SELECT COUNT(*) FROM predictions").fetchone()
        res_completed = db.execute("SELECT COUNT(*) FROM matches WHERE status='completed'").fetchone()
        return {
            "totals": {
                "users":             list(res_users.values())[0] if res_users else 0,
                "matches":           list(res_matches.values())[0] if res_matches else 0,
                "predictions":       list(res_preds.values())[0] if res_preds else 0,
                "completed_matches": list(res_completed.values())[0] if res_completed else 0,
            }
        }


# ── NOTIFY MATCH PARTICIPANTS ─────────────────────────────────────────────
@router.post("/notify/match-participants/{match_id}")
def notify_match_participants(match_id: int, admin: dict = Depends(admin_user)):
    """Send email to all users who predicted this match."""
    from ..services.emailer import send_email, queue_notification

    with get_db() as db:
        # Get match details
        match = row(db.execute("""
            SELECT m.*, ht.name AS home_team, at.name AS away_team
            FROM matches m
            JOIN teams ht ON ht.id = m.home_team_id
            JOIN teams at ON at.id = m.away_team_id
            WHERE m.id = ?
        """, (match_id,)))

        if not match:
            raise HTTPException(status_code=404, detail="Match not found")

        # Get all participants with their predictions
        participants = rows(db.execute("""
            SELECT u.id, u.email, u.name,
                   p.predicted_home_score, p.predicted_away_score,
                   p.points_awarded, p.scoring_reason
            FROM predictions p
            JOIN users u ON u.id = p.user_id
            WHERE p.match_id = ?
        """, (match_id,)))

        if not participants:
            return {
                "status": "no_participants",
                "message": "No predictions found for this match",
                "sent": 0, "skipped": 0, "failed": 0
            }

        game_label = match.get("game_no") or f"Match {match_id}"
        home       = match.get("home_team", "Home")
        away       = match.get("away_team", "Away")
        home_score = match.get("home_score")
        away_score = match.get("away_score")

        if home_score is not None and away_score is not None:
            result_line = f"Result: {home} {home_score} - {away_score} {away}"
        else:
            result_line = f"Match: {home} vs {away} (result pending)"

        sent = skipped = failed = 0

        for p in participants:
            subject = f"⚽ {game_label} Update — {home} vs {away}"
            body = (
                f"Hi {p['name']},\n\n"
                f"{result_line}\n\n"
                f"Your prediction: {p['predicted_home_score']} - {p['predicted_away_score']}\n"
                f"Points awarded:  {p['points_awarded'] or 0}\n"
                f"Result:          {p['scoring_reason'] or 'Pending'}\n\n"
                f"Check the leaderboard: https://worldcup-lv.onrender.com\n\n"
                f"Good luck!\nWorldCup Prediction Team"
            )
            try:
                status = send_email(p["email"], subject, body)
                notif_id = queue_notification(db, p["email"], subject, body, p["id"])
                db.execute(
                    "UPDATE notifications SET status = ? WHERE id = ?",
                    (status, notif_id)
                )
                if status == "sent":
                    sent += 1
                else:
                    skipped += 1
            except Exception as e:
                failed += 1
                queue_notification(db, p["email"], subject, body, p["id"])

        return {
            "status": "success",
            "match": f"{home} vs {away}",
            "sent": sent,
            "skipped": skipped,
            "failed": failed,
            "message": f"Sent: {sent}, Skipped: {skipped}, Failed: {failed}"
        }


# ── IMPORT FIFA WORLD CUP 2026 FIXTURES ──────────────────────────────────
@router.post("/import-worldcup-fixtures/{tournament_id}")
def import_worldcup_fixtures(tournament_id: int, admin: dict = Depends(admin_user)):
    from ..services.scoring import lock_time

    FIXTURES = [
        ("G01","Group A","2026-06-11T18:00:00Z","Estadio Azteca, Mexico City","Mexico","South Africa"),
        ("G02","Group A","2026-06-11T22:00:00Z","Estadio Akron, Guadalajara","South Korea","Czechia"),
        ("G03","Group A","2026-06-18T16:00:00Z","Mercedes-Benz Stadium, Atlanta","Czechia","South Africa"),
        ("G04","Group A","2026-06-19T02:00:00Z","Estadio Akron, Guadalajara","Mexico","South Korea"),
        ("G05","Group A","2026-06-25T01:00:00Z","Estadio Azteca, Mexico City","Czechia","Mexico"),
        ("G06","Group A","2026-06-25T01:00:00Z","Estadio BBVA, Monterrey","South Africa","South Korea"),
        ("G07","Group B","2026-06-12T17:00:00Z","BC Place, Vancouver","Canada","Bosnia and Herzegovina"),
        ("G08","Group B","2026-06-13T20:00:00Z","Estadio Azteca, Mexico City","Qatar","Switzerland"),
        ("G09","Group B","2026-06-18T22:00:00Z","SoFi Stadium, Los Angeles","Switzerland","Bosnia and Herzegovina"),
        ("G10","Group B","2026-06-19T01:00:00Z","BC Place, Vancouver","Canada","Qatar"),
        ("G11","Group B","2026-06-24T19:00:00Z","BC Place, Vancouver","Switzerland","Canada"),
        ("G12","Group B","2026-06-24T19:00:00Z","Lumen Field, Seattle","Bosnia and Herzegovina","Qatar"),
        ("G13","Group C","2026-06-13T20:00:00Z","Hard Rock Stadium, Miami","Brazil","Morocco"),
        ("G14","Group C","2026-06-13T20:00:00Z","Gillette Stadium, Boston","Haiti","Scotland"),
        ("G15","Group C","2026-06-19T22:00:00Z","Gillette Stadium, Boston","Scotland","Morocco"),
        ("G16","Group C","2026-06-20T01:00:00Z","Lincoln Financial Field, Philadelphia","Brazil","Haiti"),
        ("G17","Group C","2026-06-24T22:00:00Z","Hard Rock Stadium, Miami","Scotland","Brazil"),
        ("G18","Group C","2026-06-24T22:00:00Z","Mercedes-Benz Stadium, Atlanta","Morocco","Haiti"),
        ("G19","Group D","2026-06-12T19:00:00Z","SoFi Stadium, Los Angeles","USA","Paraguay"),
        ("G20","Group D","2026-06-13T20:00:00Z","Lumen Field, Seattle","Australia","Turkiye"),
        ("G21","Group D","2026-06-19T19:00:00Z","Lumen Field, Seattle","USA","Australia"),
        ("G22","Group D","2026-06-20T04:00:00Z","Levi's Stadium, San Francisco Bay Area","Turkiye","Paraguay"),
        ("G23","Group D","2026-06-26T02:00:00Z","SoFi Stadium, Los Angeles","Turkiye","USA"),
        ("G24","Group D","2026-06-26T02:00:00Z","Levi's Stadium, San Francisco Bay Area","Paraguay","Australia"),
        ("G25","Group E","2026-06-14T22:00:00Z","BMO Field, Toronto","Germany","Curacao"),
        ("G26","Group E","2026-06-14T22:00:00Z","Arrowhead Stadium, Kansas City","Ivory Coast","Ecuador"),
        ("G27","Group E","2026-06-20T20:00:00Z","BMO Field, Toronto","Germany","Ivory Coast"),
        ("G28","Group E","2026-06-21T00:00:00Z","Arrowhead Stadium, Kansas City","Ecuador","Curacao"),
        ("G29","Group E","2026-06-25T20:00:00Z","MetLife Stadium, New York/New Jersey","Ecuador","Germany"),
        ("G30","Group E","2026-06-25T20:00:00Z","Lincoln Financial Field, Philadelphia","Curacao","Ivory Coast"),
        ("G31","Group F","2026-06-15T00:00:00Z","NRG Stadium, Houston","Netherlands","Japan"),
        ("G32","Group F","2026-06-15T00:00:00Z","Estadio BBVA, Monterrey","Sweden","Tunisia"),
        ("G33","Group F","2026-06-20T17:00:00Z","NRG Stadium, Houston","Netherlands","Sweden"),
        ("G34","Group F","2026-06-21T04:00:00Z","Estadio BBVA, Monterrey","Tunisia","Japan"),
        ("G35","Group F","2026-06-25T23:00:00Z","AT&T Stadium, Dallas","Japan","Sweden"),
        ("G36","Group F","2026-06-25T23:00:00Z","Arrowhead Stadium, Kansas City","Tunisia","Netherlands"),
        ("G37","Group G","2026-06-15T00:00:00Z","SoFi Stadium, Los Angeles","Iran","New Zealand"),
        ("G38","Group G","2026-06-16T03:00:00Z","Levi's Stadium, San Francisco Bay Area","Belgium","Egypt"),
        ("G39","Group G","2026-06-21T19:00:00Z","SoFi Stadium, Los Angeles","Belgium","Iran"),
        ("G40","Group G","2026-06-22T01:00:00Z","BC Place, Vancouver","New Zealand","Egypt"),
        ("G41","Group G","2026-06-27T03:00:00Z","Lumen Field, Seattle","Egypt","Iran"),
        ("G42","Group G","2026-06-27T03:00:00Z","BC Place, Vancouver","New Zealand","Belgium"),
        ("G43","Group H","2026-06-16T00:00:00Z","Mercedes-Benz Stadium, Atlanta","Spain","Cape Verde"),
        ("G44","Group H","2026-06-16T00:00:00Z","Estadio Akron, Guadalajara","Saudi Arabia","Uruguay"),
        ("G45","Group H","2026-06-21T17:00:00Z","Mercedes-Benz Stadium, Atlanta","Spain","Saudi Arabia"),
        ("G46","Group H","2026-06-21T23:00:00Z","Hard Rock Stadium, Miami","Uruguay","Cape Verde"),
        ("G47","Group H","2026-06-27T01:00:00Z","NRG Stadium, Houston","Cape Verde","Saudi Arabia"),
        ("G48","Group H","2026-06-27T01:00:00Z","Estadio Akron, Guadalajara","Uruguay","Spain"),
        ("G49","Group I","2026-06-16T18:00:00Z","MetLife Stadium, New York/New Jersey","France","Senegal"),
        ("G50","Group I","2026-06-17T03:00:00Z","Levi's Stadium, San Francisco Bay Area","Iraq","Norway"),
        ("G51","Group I","2026-06-22T21:00:00Z","Lincoln Financial Field, Philadelphia","France","Iraq"),
        ("G52","Group I","2026-06-23T00:00:00Z","MetLife Stadium, New York/New Jersey","Norway","Senegal"),
        ("G53","Group I","2026-06-26T19:00:00Z","Gillette Stadium, Boston","Norway","France"),
        ("G54","Group I","2026-06-26T19:00:00Z","BMO Field, Toronto","Senegal","Iraq"),
        ("G55","Group J","2026-06-16T22:00:00Z","Arrowhead Stadium, Kansas City","Argentina","Algeria"),
        ("G56","Group J","2026-06-17T04:00:00Z","Levi's Stadium, San Francisco Bay Area","Austria","Jordan"),
        ("G57","Group J","2026-06-22T17:00:00Z","AT&T Stadium, Dallas","Argentina","Austria"),
        ("G58","Group J","2026-06-22T23:00:00Z","Levi's Stadium, San Francisco Bay Area","Jordan","Algeria"),
        ("G59","Group J","2026-06-28T02:00:00Z","Arrowhead Stadium, Kansas City","Algeria","Austria"),
        ("G60","Group J","2026-06-28T02:00:00Z","AT&T Stadium, Dallas","Jordan","Argentina"),
        ("G61","Group K","2026-06-17T17:00:00Z","NRG Stadium, Houston","Portugal","Congo DR"),
        ("G62","Group K","2026-06-18T02:00:00Z","Estadio Azteca, Mexico City","Uzbekistan","Colombia"),
        ("G63","Group K","2026-06-23T17:00:00Z","NRG Stadium, Houston","Portugal","Uzbekistan"),
        ("G64","Group K","2026-06-24T02:00:00Z","Estadio Akron, Guadalajara","Colombia","Congo DR"),
        ("G65","Group K","2026-06-27T23:30:00Z","Hard Rock Stadium, Miami","Colombia","Portugal"),
        ("G66","Group K","2026-06-27T23:30:00Z","Mercedes-Benz Stadium, Atlanta","Congo DR","Uzbekistan"),
        ("G67","Group L","2026-06-17T20:00:00Z","AT&T Stadium, Dallas","England","Croatia"),
        ("G68","Group L","2026-06-17T23:00:00Z","BMO Field, Toronto","Ghana","Panama"),
        ("G69","Group L","2026-06-23T20:00:00Z","Gillette Stadium, Boston","England","Ghana"),
        ("G70","Group L","2026-06-23T23:00:00Z","BMO Field, Toronto","Panama","Croatia"),
        ("G71","Group L","2026-06-28T02:00:00Z","Lincoln Financial Field, Philadelphia","Croatia","Ghana"),
        ("G72","Group L","2026-06-28T02:00:00Z","MetLife Stadium, New York/New Jersey","Panama","England"),
        ("R32-1","Round of 32","2026-06-28T22:00:00Z","MetLife Stadium, New York/NJ","TBD","TBD"),
        ("R32-2","Round of 32","2026-06-29T01:00:00Z","AT&T Stadium, Dallas","TBD","TBD"),
        ("R32-3","Round of 32","2026-06-29T19:00:00Z","Rose Bowl, Los Angeles","TBD","TBD"),
        ("R32-4","Round of 32","2026-06-29T22:00:00Z","NRG Stadium, Houston","TBD","TBD"),
        ("R32-5","Round of 32","2026-06-30T01:00:00Z","BC Place Stadium, Vancouver","TBD","TBD"),
        ("R32-6","Round of 32","2026-06-30T19:00:00Z","Hard Rock Stadium, Miami","TBD","TBD"),
        ("R32-7","Round of 32","2026-06-30T22:00:00Z","SoFi Stadium, Los Angeles","TBD","TBD"),
        ("R32-8","Round of 32","2026-07-01T01:00:00Z","Lincoln Financial Field, Philadelphia","TBD","TBD"),
        ("R32-9","Round of 32","2026-07-01T19:00:00Z","Estadio Azteca, Mexico City","TBD","TBD"),
        ("R32-10","Round of 32","2026-07-01T22:00:00Z","Arrowhead Stadium, Kansas City","TBD","TBD"),
        ("R32-11","Round of 32","2026-07-02T01:00:00Z","Gillette Stadium, Boston","TBD","TBD"),
        ("R32-12","Round of 32","2026-07-02T19:00:00Z","Levi's Stadium, San Francisco","TBD","TBD"),
        ("R32-13","Round of 32","2026-07-02T22:00:00Z","Estadio BBVA, Monterrey","TBD","TBD"),
        ("R32-14","Round of 32","2026-07-03T01:00:00Z","AT&T Stadium, Dallas","TBD","TBD"),
        ("R32-15","Round of 32","2026-07-03T19:00:00Z","MetLife Stadium, New York/NJ","TBD","TBD"),
        ("R32-16","Round of 32","2026-07-03T22:00:00Z","Rose Bowl, Los Angeles","TBD","TBD"),
        ("R16-1","Round of 16","2026-07-04T17:00:00Z","NRG Stadium, Houston","TBD","TBD"),
        ("R16-2","Round of 16","2026-07-04T21:00:00Z","Lincoln Financial Field, Philadelphia","TBD","TBD"),
        ("R16-3","Round of 16","2026-07-05T20:00:00Z","MetLife Stadium, New York/NJ","TBD","TBD"),
        ("R16-4","Round of 16","2026-07-06T00:00:00Z","Estadio Azteca, Mexico City","TBD","TBD"),
        ("R16-5","Round of 16","2026-07-06T19:00:00Z","AT&T Stadium, Dallas","TBD","TBD"),
        ("R16-6","Round of 16","2026-07-07T00:00:00Z","Lumen Field, Seattle","TBD","TBD"),
        ("R16-7","Round of 16","2026-07-07T16:00:00Z","Mercedes-Benz Stadium, Atlanta","TBD","TBD"),
        ("R16-8","Round of 16","2026-07-07T20:00:00Z","BC Place Stadium, Vancouver","TBD","TBD"),
        ("QF-1","Quarter-Final","2026-07-09T20:00:00Z","Gillette Stadium, Boston","TBD","TBD"),
        ("QF-2","Quarter-Final","2026-07-10T19:00:00Z","SoFi Stadium, Los Angeles","TBD","TBD"),
        ("QF-3","Quarter-Final","2026-07-11T21:00:00Z","Hard Rock Stadium, Miami","TBD","TBD"),
        ("QF-4","Quarter-Final","2026-07-12T01:00:00Z","Arrowhead Stadium, Kansas City","TBD","TBD"),
        ("SF-1","Semi-Final","2026-07-14T19:00:00Z","AT&T Stadium, Dallas","TBD","TBD"),
        ("SF-2","Semi-Final","2026-07-15T19:00:00Z","Mercedes-Benz Stadium, Atlanta","TBD","TBD"),
        ("3P-1","3rd Place","2026-07-18T21:00:00Z","Hard Rock Stadium, Miami","TBD","TBD"),
        ("FINAL","Final","2026-07-19T19:00:00Z","MetLife Stadium, New York/New Jersey","TBD","TBD"),
    ]

    KNOWN_FLAGS = {
        "Mexico":"\U0001F1F2\U0001F1FD","South Korea":"\U0001F1F0\U0001F1F7",
        "Czechia":"\U0001F1E8\U0001F1FF","South Africa":"\U0001F1FF\U0001F1E6",
        "Canada":"\U0001F1E8\U0001F1E6","Bosnia and Herzegovina":"\U0001F1E7\U0001F1E6",
        "Qatar":"\U0001F1F6\U0001F1E6","Switzerland":"\U0001F1E8\U0001F1ED",
        "Brazil":"\U0001F1E7\U0001F1F7","Morocco":"\U0001F1F2\U0001F1E6",
        "Haiti":"\U0001F1ED\U0001F1F9","Scotland":"\U0001F3F4",
        "USA":"\U0001F1FA\U0001F1F8","Paraguay":"\U0001F1F5\U0001F1FE",
        "Australia":"\U0001F1E6\U0001F1FA","Turkiye":"\U0001F1F9\U0001F1F7",
        "Germany":"\U0001F1E9\U0001F1EA","Curacao":"\U0001F1E8\U0001F1FC",
        "Ivory Coast":"\U0001F1E8\U0001F1EE","Ecuador":"\U0001F1EA\U0001F1E8",
        "Netherlands":"\U0001F1F3\U0001F1F1","Japan":"\U0001F1EF\U0001F1F5",
        "Sweden":"\U0001F1F8\U0001F1EA","Tunisia":"\U0001F1F9\U0001F1F3",
        "Belgium":"\U0001F1E7\U0001F1EA","Egypt":"\U0001F1EA\U0001F1EC",
        "Iran":"\U0001F1EE\U0001F1F7","New Zealand":"\U0001F1F3\U0001F1FF",
        "Spain":"\U0001F1EA\U0001F1F8","Cape Verde":"\U0001F1E8\U0001F1FB",
        "Saudi Arabia":"\U0001F1F8\U0001F1E6","Uruguay":"\U0001F1FA\U0001F1FE",
        "France":"\U0001F1EB\U0001F1F7","Senegal":"\U0001F1F8\U0001F1F3",
        "Iraq":"\U0001F1EE\U0001F1F6","Norway":"\U0001F1F3\U0001F1F4",
        "Argentina":"\U0001F1E6\U0001F1F7","Algeria":"\U0001F1E9\U0001F1FF",
        "Austria":"\U0001F1E6\U0001F1F9","Jordan":"\U0001F1EF\U0001F1F4",
        "Portugal":"\U0001F1F5\U0001F1F9","Congo DR":"\U0001F1E8\U0001F1E9",
        "Uzbekistan":"\U0001F1FA\U0001F1FF","Colombia":"\U0001F1E8\U0001F1F4",
        "England":"\U0001F3F4\U000E0067\U000E0062\U000E0065\U000E006E\U000E0067\U000E007F",
        "Croatia":"\U0001F1ED\U0001F1F7","Ghana":"\U0001F1EC\U0001F1ED","Panama":"\U0001F1F5\U0001F1E6",
    }

    with get_db() as db:
        tournament = row(db.execute("SELECT id FROM tournaments WHERE id = ?", (tournament_id,)))
        if not tournament:
            raise HTTPException(status_code=404, detail="Tournament not found")

        def get_or_create_team(name: str) -> int:
            existing = row(db.execute(
                "SELECT id FROM teams WHERE tournament_id=? AND name=?",
                (tournament_id, name)
            ))
            if existing:
                return existing["id"]
            cur = db.execute(
                "INSERT INTO teams (tournament_id, name, country, flag, ranking, strength_score) VALUES (?,?,?,?,50,50)",
                (tournament_id, name, name, KNOWN_FLAGS.get(name, "\U0001F3C6"))
            )
            return cur.lastrowid

        imported = skipped = 0
        errors = []

        for game_no, round_name, date, stadium, home, away in FIXTURES:
            exists = row(db.execute(
                "SELECT id FROM matches WHERE tournament_id=? AND game_no=?",
                (tournament_id, game_no)
            ))
            if exists:
                skipped += 1
                continue

            home_name = home if home != "TBD" else f"TBD ({game_no} Home)"
            away_name = away if away != "TBD" else f"TBD ({game_no} Away)"

            try:
                home_id = get_or_create_team(home_name)
                away_id = get_or_create_team(away_name)
                if home_id == away_id:
                    errors.append(f"{game_no}: same team for home/away, skipped")
                    skipped += 1
                    continue
                db.execute(
                    """
                    INSERT INTO matches
                        (tournament_id, game_no, sport, round, match_date, lock_at,
                         stadium, home_team_id, away_team_id, status, predictions_open)
                    VALUES (?,?,?,?,?,?,?,?,?,'scheduled',0)
                    """,
                    (tournament_id, game_no, "FIFA World Cup", round_name,
                     date, lock_time(date), stadium, home_id, away_id)
                )
                imported += 1
            except Exception as e:
                errors.append(f"{game_no}: {e}")
                skipped += 1

    return {
        "imported": imported, "skipped": skipped,
        "total": imported + skipped, "errors": errors,
        "source": "FIFA World Cup 2026 Official Fixtures",
    }


# ── UPLOAD SCHEDULE ───────────────────────────────────────────────────────
@router.post("/upload-schedule/{tournament_id}")
async def upload_schedule(
    tournament_id: int,
    file: UploadFile = File(...),
    admin: dict = Depends(admin_user),
):
    from ..services.scoring import lock_time
    import io
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    required = {"team_a", "team_b", "match_date"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}")

    def col(row_data, name):
        try:
            idx = headers.index(name)
            return str(row_data[idx].value).strip() if row_data[idx].value is not None else ""
        except (ValueError, IndexError):
            return ""

    imported = skipped = 0
    errors = []

    with get_db() as db:
        tournament = row(db.execute("SELECT id FROM tournaments WHERE id=?", (tournament_id,)))
        if not tournament:
            raise HTTPException(status_code=404, detail="Tournament not found")

        def get_or_create_team(name: str) -> int:
            if not name:
                raise ValueError("Team name is empty")
            existing = row(db.execute("SELECT id FROM teams WHERE tournament_id=? AND name=?", (tournament_id, name)))
            if existing:
                return existing["id"]
            cur = db.execute(
                "INSERT INTO teams (tournament_id, name, country, flag, ranking, strength_score) VALUES (?,?,?,?,50,50)",
                (tournament_id, name, name, "🏆")
            )
            return cur.lastrowid

        for i, row_data in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                team_a     = col(row_data, "team_a")
                team_b     = col(row_data, "team_b")
                date_val   = col(row_data, "match_date")
                game_no    = col(row_data, "game_no") or f"G{i}"
                venue      = col(row_data, "venue") or "TBD"
                sport      = col(row_data, "sport") or "FIFA World Cup"
                round_name = col(row_data, "round") or "Group Stage"

                if not team_a or not team_b or not date_val:
                    skipped += 1
                    continue

                exists = row(db.execute(
                    "SELECT id FROM matches WHERE tournament_id=? AND game_no=?",
                    (tournament_id, game_no)
                ))
                if exists:
                    skipped += 1
                    continue

                home_id = get_or_create_team(team_a)
                away_id = get_or_create_team(team_b)
                db.execute(
                    """
                    INSERT INTO matches
                        (tournament_id, game_no, sport, round, match_date, lock_at,
                         stadium, home_team_id, away_team_id, status, predictions_open)
                    VALUES (?,?,?,?,?,?,?,?,?,'scheduled',0)
                    """,
                    (tournament_id, game_no, sport, round_name,
                     date_val, lock_time(date_val), venue, home_id, away_id)
                )
                imported += 1
            except Exception as e:
                errors.append(f"Row {i}: {e}")

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ── GENERATE EXPORTS ──────────────────────────────────────────────────────
@router.post("/exports/{tournament_id}")
def generate_exports(tournament_id: int, admin: dict = Depends(admin_user)):
    with get_db() as db:
        try:
            preds = rows(db.execute(
                """
                SELECT u.name AS player, m.game_no,
                       ht.name AS home_team, at.name AS away_team,
                       p.predicted_home_score, p.predicted_away_score,
                       m.home_score, m.away_score, m.status,
                       p.points_awarded, p.scoring_reason, p.created_at
                FROM predictions p
                JOIN users u ON u.id = p.user_id
                JOIN matches m ON m.id = p.match_id
                JOIN teams ht ON ht.id = m.home_team_id
                JOIN teams at ON at.id = m.away_team_id
                WHERE m.tournament_id = ?
                """, (tournament_id,)
            ))
            return {
                "status": "success",
                "message": f"Data compiled for tournament {tournament_id}.",
                "records_found": len(preds) if preds else 0
            }
        except Exception as e:
            return {"status": "error", "detail": str(e)}
